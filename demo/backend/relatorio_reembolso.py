#%%
import openpyxl
from io import BytesIO

def relatorio(datas, valor, empresa, local_ida, local_volta, caminho_arquivo):

    dic_empresa = {
    'Transporte': ['UBER', '99', 'TAXI']
    }
    dic_contabil = {
        'Transporte': {'Resumo contábil': 'Reembolso de Condução e Transporte PROADI SUS', 'Conta contábil': 'PR50060003'}
    }

    # 1. Carregar o arquivo existente
    planilha_aberta = openpyxl.load_workbook(caminho_arquivo)

    # 2. Selecionar a aba (pelo nome ou a que estiver ativa)
    # aba = planilha_aberta.active
    aba = planilha_aberta['MATRIZ - Relat Despesas']

    linha_inicial = 30 # Linha de inicio do preenchimento
    # Enquanto a célula na coluna B (data) não estiver vazia, incrementa a linha
    while aba.cell(row=linha_inicial, column=21).value is not None:
        linha_inicial += 1

    coluna_data_b = 2 # Coluna de preenchimento das datas
    coluna_data_au = 47 # Coluna de preenchimento dos valores
    coluna_empresa_u = 21 # Coluna de preenchimento dos empresas
    coluna_classificacao_despesa_ao = 41 # Coluna da classificação da despesa
    col_cod_contabil_ar = 44 # Coluna do código contábil
    col_descricao_aa = 27 # Coluna da descrição

    # 3. Inserir dados em células específicas
    for i, data in enumerate(datas):
            aba.cell(row=linha_inicial + i, column=coluna_data_b, value=data)

    for i, valor in enumerate(valor):
            aba.cell(row=linha_inicial + i, column=coluna_data_au, value=valor)

    for i, empresa in enumerate(empresa):
        aba.cell(row=linha_inicial + i, column=coluna_empresa_u, value=empresa)

        # 1. Descobrir a categoria (Transporte)
        categoria_encontrada = None
        for cat, empresas in dic_empresa.items():
            if empresa in empresas:
                categoria_encontrada = cat
                break

        # 2. Buscar no dicionário contábil usando a categoria achada
        if categoria_encontrada:
            info_contabil = dic_contabil.get(categoria_encontrada)
            resumo = info_contabil['Resumo contábil']
            conta = info_contabil['Conta contábil']

        aba.cell(row=linha_inicial + i, column=coluna_classificacao_despesa_ao, value=resumo)
        aba.cell(row=linha_inicial + i, column=col_cod_contabil_ar, value=conta)

        if local_ida:
            for i, local in enumerate(local_ida):
                aba.cell(row=linha_inicial + i, column=col_descricao_aa, value=empresa + ' ' + local + ' - ' + local_volta[i])

    # 3. Salvar o resultado em memória (BytesIO)
    output = BytesIO()
    planilha_aberta.save(output)
    output.seek(0)  # Volta para o início do arquivo para leitura
    return output
